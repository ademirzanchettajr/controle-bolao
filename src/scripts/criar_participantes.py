#!/usr/bin/env python3
"""
Script de criação de participantes para o Sistema de Controle de Bolão.

Este script cria a estrutura de diretórios e arquivos para participantes
de um campeonato, incluindo leitura de listas de nomes de arquivos texto
ou planilhas Excel, normalização de nomes e validação de duplicados.
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Set, Tuple
import random
import string

# Adicionar o diretório pai ao path para imports
sys.path.append(str(Path(__file__).parent.parent))

from config import (
    CAMPEONATOS_DIR, 
    ARQUIVO_PALPITES
)
from utils.normalizacao import normalizar_nome_participante
from utils.validacao import validar_estrutura_palpites


def gerar_codigo_participante() -> str:
    """
    Gera código único de 4 dígitos para o participante.
    
    Returns:
        String com código único alfanumérico
    """
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))


def ler_nomes_arquivo_texto(caminho_arquivo: Path) -> List[str]:
    """
    Lê lista de nomes de arquivo texto (um nome por linha).
    
    Args:
        caminho_arquivo: Path para o arquivo texto
        
    Returns:
        Lista de nomes extraídos do arquivo
        
    Raises:
        OSError: Se não conseguir ler o arquivo
        ValueError: Se o arquivo estiver vazio ou inválido
    """
    try:
        if not caminho_arquivo.exists():
            raise OSError(f"Arquivo não encontrado: {caminho_arquivo}")
        
        with open(caminho_arquivo, 'r', encoding='utf-8') as f:
            linhas = f.readlines()
        
        # Filtrar linhas vazias e remover espaços
        nomes = []
        for i, linha in enumerate(linhas, 1):
            nome = linha.strip()
            if nome:  # Ignora linhas vazias
                nomes.append(nome)
        
        if not nomes:
            raise ValueError(f"Arquivo '{caminho_arquivo}' não contém nomes válidos")
        
        return nomes
        
    except UnicodeDecodeError:
        raise OSError(f"Erro de codificação ao ler arquivo '{caminho_arquivo}'. Verifique se está em UTF-8")
    except Exception as e:
        raise OSError(f"Erro ao ler arquivo '{caminho_arquivo}': {str(e)}")


def ler_nomes_planilha_excel(caminho_arquivo: Path, nome_coluna: str = "Nome") -> List[str]:
    """
    Lê lista de nomes de planilha Excel.
    
    Args:
        caminho_arquivo: Path para o arquivo Excel
        nome_coluna: Nome da coluna que contém os nomes
        
    Returns:
        Lista de nomes extraídos da planilha
        
    Raises:
        OSError: Se não conseguir ler o arquivo
        ValueError: Se a coluna não for encontrada ou estiver vazia
        ImportError: Se openpyxl não estiver instalado
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Biblioteca 'openpyxl' não encontrada. Instale com: pip install openpyxl")
    
    try:
        if not caminho_arquivo.exists():
            raise OSError(f"Arquivo não encontrado: {caminho_arquivo}")
        
        # Carregar planilha
        workbook = openpyxl.load_workbook(caminho_arquivo)
        sheet = workbook.active
        
        # Encontrar coluna com os nomes
        coluna_nomes = None
        for col in range(1, sheet.max_column + 1):
            valor_cabecalho = sheet.cell(row=1, column=col).value
            if valor_cabecalho and str(valor_cabecalho).strip().lower() == nome_coluna.lower():
                coluna_nomes = col
                break
        
        if coluna_nomes is None:
            colunas_disponiveis = []
            for col in range(1, sheet.max_column + 1):
                valor = sheet.cell(row=1, column=col).value
                if valor:
                    colunas_disponiveis.append(str(valor))
            
            raise ValueError(f"Coluna '{nome_coluna}' não encontrada. Colunas disponíveis: {', '.join(colunas_disponiveis)}")
        
        # Extrair nomes da coluna (começando da linha 2, pulando cabeçalho)
        nomes = []
        for row in range(2, sheet.max_row + 1):
            valor = sheet.cell(row=row, column=coluna_nomes).value
            if valor:
                nome = str(valor).strip()
                if nome:
                    nomes.append(nome)
        
        if not nomes:
            raise ValueError(f"Coluna '{nome_coluna}' não contém nomes válidos")
        
        return nomes
        
    except Exception as e:
        if isinstance(e, (OSError, ValueError, ImportError)):
            raise
        else:
            raise OSError(f"Erro ao ler planilha Excel '{caminho_arquivo}': {str(e)}")


def obter_dados_campeonato_existente(caminho_campeonato: Path) -> Optional[Tuple[str, str]]:
    """
    Obtém dados do campeonato existente a partir do arquivo tabela.json.
    
    Args:
        caminho_campeonato: Path para o diretório do campeonato
        
    Returns:
        Tupla (nome_campeonato, temporada) ou None se não conseguir obter
    """
    try:
        caminho_tabela = caminho_campeonato / "Tabela" / "tabela.json"
        
        if not caminho_tabela.exists():
            return None
            
        with open(caminho_tabela, 'r', encoding='utf-8') as f:
            dados_tabela = json.load(f)
            
        nome = dados_tabela.get("campeonato")
        temporada = dados_tabela.get("temporada")
        
        if nome and temporada:
            return (nome, temporada)
            
        return None
        
    except Exception:
        return None


def verificar_campeonato_existe(nome_campeonato: str) -> Tuple[bool, Path, Optional[Tuple[str, str]]]:
    """
    Verifica se o campeonato existe e retorna informações.
    
    Args:
        nome_campeonato: Nome normalizado do campeonato
        
    Returns:
        Tupla (existe: bool, caminho: Path, dados: tuple ou None)
    """
    caminho_campeonato = CAMPEONATOS_DIR / nome_campeonato
    
    if not caminho_campeonato.exists():
        return (False, caminho_campeonato, None)
    
    # Verificar se é um diretório válido de campeonato
    if not caminho_campeonato.is_dir():
        return (False, caminho_campeonato, None)
    
    # Tentar obter dados do campeonato
    dados = obter_dados_campeonato_existente(caminho_campeonato)
    
    return (True, caminho_campeonato, dados)


def obter_participantes_existentes(caminho_campeonato: Path) -> Set[str]:
    """
    Obtém lista de participantes já existentes no campeonato.
    
    Args:
        caminho_campeonato: Path para o diretório do campeonato
        
    Returns:
        Set com nomes normalizados dos participantes existentes
    """
    participantes_existentes = set()
    
    caminho_participantes = caminho_campeonato / "Participantes"
    
    if caminho_participantes.exists():
        try:
            for item in caminho_participantes.iterdir():
                if item.is_dir():
                    participantes_existentes.add(item.name)
        except Exception:
            pass  # Ignora erros de acesso
    
    return participantes_existentes


def criar_estrutura_basica_palpites(nome_participante: str, codigo_participante: str, 
                                   nome_campeonato: str, temporada: str) -> dict:
    """
    Cria estrutura básica do arquivo palpites.json para um participante.
    
    Args:
        nome_participante: Nome original do participante
        codigo_participante: Código único do participante
        nome_campeonato: Nome do campeonato
        temporada: Temporada do campeonato
        
    Returns:
        Dicionário com estrutura básica dos palpites
    """
    return {
        "apostador": nome_participante,
        "codigo_apostador": codigo_participante,
        "campeonato": nome_campeonato,
        "temporada": temporada,
        "palpites": []
    }


def criar_diretorio_participante(caminho_participantes: Path, nome_normalizado: str) -> Path:
    """
    Cria o diretório do participante.
    
    Args:
        caminho_participantes: Path para o diretório de participantes
        nome_normalizado: Nome normalizado do participante
        
    Returns:
        Path para o diretório criado
        
    Raises:
        OSError: Se não conseguir criar o diretório
    """
    try:
        # Garantir que o diretório Participantes existe
        caminho_participantes.mkdir(parents=True, exist_ok=True)
        
        # Criar diretório do participante
        caminho_participante = caminho_participantes / nome_normalizado
        caminho_participante.mkdir(parents=True, exist_ok=True)
        
        return caminho_participante
        
    except Exception as e:
        raise OSError(f"Erro ao criar diretório do participante '{nome_normalizado}': {str(e)}")


def criar_arquivo_palpites_vazio(caminho_participante: Path, nome_participante: str, 
                                codigo_participante: str, nome_campeonato: str, 
                                temporada: str) -> None:
    """
    Cria o arquivo palpites.json vazio para o participante.
    
    Args:
        caminho_participante: Path para o diretório do participante
        nome_participante: Nome original do participante
        codigo_participante: Código único do participante
        nome_campeonato: Nome do campeonato
        temporada: Temporada do campeonato
        
    Raises:
        OSError: Se não conseguir criar o arquivo
        ValueError: Se a estrutura gerada for inválida
    """
    try:
        # Criar estrutura básica
        estrutura_palpites = criar_estrutura_basica_palpites(
            nome_participante, codigo_participante, nome_campeonato, temporada
        )
        
        # Validar estrutura antes de salvar
        valido, erros = validar_estrutura_palpites(estrutura_palpites)
        if not valido:
            raise ValueError(f"Estrutura dos palpites inválida: {'; '.join(erros)}")
        
        # Escrever arquivo
        caminho_palpites = caminho_participante / ARQUIVO_PALPITES
        with open(caminho_palpites, 'w', encoding='utf-8') as f:
            json.dump(estrutura_palpites, f, indent=2, ensure_ascii=False)
            
    except Exception as e:
        if isinstance(e, (OSError, ValueError)):
            raise
        else:
            raise OSError(f"Erro ao criar arquivo de palpites: {str(e)}")


def confirmar_operacao(mensagem: str) -> bool:
    """
    Solicita confirmação do usuário para operações críticas.
    
    Args:
        mensagem: Mensagem a ser exibida
        
    Returns:
        True se o usuário confirmar, False caso contrário
    """
    resposta = input(f"{mensagem} (s/n): ").strip().lower()
    return resposta in ['s', 'sim', 'y', 'yes']


def processar_lista_participantes(nomes_originais: List[str]) -> Tuple[List[Tuple[str, str]], List[str]]:
    """
    Processa lista de nomes, normalizando e detectando duplicados.
    
    Args:
        nomes_originais: Lista de nomes originais
        
    Returns:
        Tupla (lista_processada, lista_duplicados)
        onde lista_processada contém tuplas (nome_original, nome_normalizado)
        e lista_duplicados contém nomes que resultaram em duplicados após normalização
    """
    nomes_processados = []
    nomes_normalizados_vistos = set()
    duplicados = []
    
    for nome_original in nomes_originais:
        nome_normalizado = normalizar_nome_participante(nome_original)
        
        if not nome_normalizado:
            print(f"Aviso: Nome '{nome_original}' resultou em string vazia após normalização - ignorado")
            continue
        
        if nome_normalizado in nomes_normalizados_vistos:
            duplicados.append(nome_original)
        else:
            nomes_normalizados_vistos.add(nome_normalizado)
            nomes_processados.append((nome_original, nome_normalizado))
    
    return nomes_processados, duplicados


def criar_participantes(nome_campeonato: str, nomes_participantes: List[str], 
                       forcar: bool = False) -> bool:
    """
    Função principal para criar participantes de um campeonato.
    
    Args:
        nome_campeonato: Nome normalizado do campeonato
        nomes_participantes: Lista de nomes dos participantes
        forcar: Se True, não solicita confirmação para operações
        
    Returns:
        True se os participantes foram criados com sucesso, False caso contrário
    """
    try:
        # Verificar se campeonato existe
        existe, caminho_campeonato, dados_campeonato = verificar_campeonato_existe(nome_campeonato)
        
        if not existe:
            print(f"Erro: Campeonato '{nome_campeonato}' não encontrado")
            print(f"Verifique se o diretório existe em: {caminho_campeonato}")
            return False
        
        # Obter dados do campeonato
        if dados_campeonato:
            nome_original_campeonato, temporada = dados_campeonato
            print(f"Campeonato encontrado: {nome_original_campeonato} ({temporada})")
        else:
            print(f"Aviso: Não foi possível obter dados do campeonato de tabela.json")
            print("Será necessário fornecer nome e temporada manualmente")
            
            nome_original_campeonato = input("Digite o nome do campeonato: ").strip()
            temporada = input("Digite a temporada: ").strip()
            
            if not nome_original_campeonato or not temporada:
                print("Erro: Nome e temporada são obrigatórios")
                return False
        
        # Processar lista de participantes
        print(f"Processando {len(nomes_participantes)} nomes de participantes...")
        
        nomes_processados, duplicados = processar_lista_participantes(nomes_participantes)
        
        if duplicados:
            print(f"Aviso: {len(duplicados)} nomes resultaram em duplicados após normalização:")
            for dup in duplicados:
                print(f"  - {dup}")
            
            if not forcar:
                if not confirmar_operacao("Deseja continuar mesmo assim?"):
                    print("Operação cancelada pelo usuário")
                    return False
        
        if not nomes_processados:
            print("Erro: Nenhum nome válido para processar")
            return False
        
        # Verificar participantes já existentes
        participantes_existentes = obter_participantes_existentes(caminho_campeonato)
        
        novos_participantes = []
        participantes_ja_existem = []
        
        for nome_original, nome_normalizado in nomes_processados:
            if nome_normalizado in participantes_existentes:
                participantes_ja_existem.append((nome_original, nome_normalizado))
            else:
                novos_participantes.append((nome_original, nome_normalizado))
        
        if participantes_ja_existem:
            print(f"Aviso: {len(participantes_ja_existem)} participantes já existem:")
            for nome_orig, nome_norm in participantes_ja_existem:
                print(f"  - {nome_orig} (diretório: {nome_norm})")
            
            if not forcar:
                if not confirmar_operacao("Deseja continuar criando apenas os novos participantes?"):
                    print("Operação cancelada pelo usuário")
                    return False
        
        if not novos_participantes:
            print("Todos os participantes já existem. Nada a fazer.")
            return True
        
        print(f"Criando {len(novos_participantes)} novos participantes...")
        
        # Criar estrutura para cada participante
        caminho_participantes = caminho_campeonato / "Participantes"
        participantes_criados = 0
        
        for nome_original, nome_normalizado in novos_participantes:
            try:
                # Gerar código único
                codigo_participante = gerar_codigo_participante()
                
                print(f"Criando participante: {nome_original} (diretório: {nome_normalizado}, código: {codigo_participante})")
                
                # Criar diretório
                caminho_participante = criar_diretorio_participante(caminho_participantes, nome_normalizado)
                
                # Criar arquivo palpites.json
                criar_arquivo_palpites_vazio(
                    caminho_participante, nome_original, codigo_participante,
                    nome_original_campeonato, temporada
                )
                
                participantes_criados += 1
                print(f"✓ Participante '{nome_original}' criado com sucesso")
                
            except Exception as e:
                print(f"✗ Erro ao criar participante '{nome_original}': {str(e)}")
                continue
        
        # Resumo final
        print(f"\n🎉 Processo concluído!")
        print(f"✓ {participantes_criados} participantes criados com sucesso")
        
        if participantes_ja_existem:
            print(f"⚠ {len(participantes_ja_existem)} participantes já existiam")
        
        if duplicados:
            print(f"⚠ {len(duplicados)} nomes duplicados ignorados")
        
        # Sugerir próximos passos
        print("\n📋 Próximos passos sugeridos:")
        print(f"1. Execute: python importar_tabela.py --campeonato '{nome_campeonato}' --arquivo jogos.txt")
        print(f"2. Execute: python importar_palpites.py --campeonato '{nome_campeonato}' --arquivo palpite.txt")
        print(f"3. Execute: python processar_resultados.py --campeonato '{nome_campeonato}' --rodada 1 --teste")
        
        return True
        
    except OSError as e:
        print(f"Erro de sistema: {str(e)}")
        return False
    except ValueError as e:
        print(f"Erro de validação: {str(e)}")
        return False
    except Exception as e:
        print(f"Erro inesperado: {str(e)}")
        return False


def main():
    """Função principal do script."""
    parser = argparse.ArgumentParser(
        description="Cria estrutura de participantes para um campeonato",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos de uso:
  python criar_participantes.py --campeonato "brasileirao-2025" --arquivo "participantes.txt"
  python criar_participantes.py --campeonato "copa-do-brasil" --excel "participantes.xlsx"
  python criar_participantes.py --campeonato "paulistao" --excel "lista.xlsx" --coluna "Apostadores"
        """
    )
    
    parser.add_argument(
        '--campeonato',
        required=True,
        help='Nome do campeonato (deve corresponder ao nome do diretório)'
    )
    
    # Grupo mutuamente exclusivo para fonte dos nomes
    fonte_group = parser.add_mutually_exclusive_group(required=True)
    
    fonte_group.add_argument(
        '--arquivo',
        help='Arquivo texto com lista de nomes (um por linha)'
    )
    
    fonte_group.add_argument(
        '--excel',
        help='Planilha Excel com coluna de nomes'
    )
    
    parser.add_argument(
        '--coluna',
        default='Nome',
        help='Nome da coluna no Excel que contém os nomes (padrão: "Nome")'
    )
    
    parser.add_argument(
        '--forcar',
        action='store_true',
        help='Força criação sem confirmação, mesmo com duplicados ou participantes existentes'
    )
    
    args = parser.parse_args()
    
    # Validar argumentos
    if not args.campeonato.strip():
        print("Erro: Nome do campeonato não pode estar vazio")
        sys.exit(1)
    
    if not args.coluna.strip():
        print("Erro: Nome da coluna não pode estar vazio")
        sys.exit(1)
    
    # Determinar fonte dos nomes e ler dados
    try:
        if args.arquivo:
            caminho_arquivo = Path(args.arquivo)
            print(f"Lendo nomes do arquivo texto: {caminho_arquivo}")
            nomes_participantes = ler_nomes_arquivo_texto(caminho_arquivo)
        
        elif args.excel:
            caminho_excel = Path(args.excel)
            print(f"Lendo nomes da planilha Excel: {caminho_excel} (coluna: {args.coluna})")
            nomes_participantes = ler_nomes_planilha_excel(caminho_excel, args.coluna)
        
        else:
            print("Erro: Deve especificar --arquivo ou --excel")
            sys.exit(1)
        
        print(f"✓ {len(nomes_participantes)} nomes lidos com sucesso")
        
    except (OSError, ValueError, ImportError) as e:
        print(f"Erro ao ler fonte de dados: {str(e)}")
        sys.exit(1)
    
    # Executar criação de participantes
    sucesso = criar_participantes(
        nome_campeonato=args.campeonato.strip(),
        nomes_participantes=nomes_participantes,
        forcar=args.forcar
    )
    
    if sucesso:
        print("\n✅ Script executado com sucesso!")
        sys.exit(0)
    else:
        print("\n❌ Falha na execução do script")
        sys.exit(1)


if __name__ == "__main__":
    main()