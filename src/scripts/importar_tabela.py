#!/usr/bin/env python3
"""
Script de importação de tabela para o Sistema de Controle de Bolão.

Este script importa jogos de arquivos externos (texto ou Excel) para o arquivo tabela.json,
organizando por rodadas, normalizando nomes de times e convertendo datas para ISO 8601.
"""

import argparse
import json
import sys
import re
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

# Adicionar o diretório pai ao path para imports
sys.path.append(str(Path(__file__).parent.parent))

from config import (
    CAMPEONATOS_DIR, 
    ARQUIVO_TABELA,
    FORMATOS_DATA
)
from utils.normalizacao import normalizar_nome_time, encontrar_time_similar
from utils.validacao import validar_estrutura_tabela, validar_data, validar_placar
from utils.parser import extrair_rodada


def gerar_id_jogo(contador: int) -> str:
    """
    Gera ID único para um jogo.
    
    Args:
        contador: Número sequencial do jogo
        
    Returns:
        ID único no formato "jogo-XXX"
    """
    return f"jogo-{contador:03d}"


def converter_data_iso8601(data_str: str, hora_str: Optional[str] = None) -> str:
    """
    Converte data para formato ISO 8601.
    
    Args:
        data_str: String com data
        hora_str: String com hora (opcional)
        
    Returns:
        Data no formato ISO 8601
        
    Raises:
        ValueError: Se não conseguir converter a data
    """
    # Combinar data e hora se fornecidas separadamente
    if hora_str:
        data_completa = f"{data_str} {hora_str}"
    else:
        data_completa = data_str
    
    # Tentar cada formato configurado
    for formato in FORMATOS_DATA:
        try:
            dt = datetime.strptime(data_completa, formato)
            return dt.strftime("%Y-%m-%dT%H:%M:%SZ")
        except ValueError:
            continue
    
    # Se nenhum formato funcionou, tentar formatos adicionais comuns
    formatos_extras = [
        "%d/%m/%Y",      # "25/12/2024"
        "%Y-%m-%d",      # "2024-12-25"
        "%d-%m-%Y",      # "25-12-2024"
        "%d/%m/%y",      # "25/12/24"
    ]
    
    for formato in formatos_extras:
        try:
            dt = datetime.strptime(data_str, formato)
            # Se só tem data, assumir horário padrão (16:00)
            if hora_str:
                # Tentar parsear hora separadamente
                try:
                    hora_dt = datetime.strptime(hora_str, "%H:%M")
                    dt = dt.replace(hour=hora_dt.hour, minute=hora_dt.minute)
                except ValueError:
                    # Se não conseguir parsear hora, usar 16:00
                    dt = dt.replace(hour=16, minute=0)
            else:
                dt = dt.replace(hour=16, minute=0)
            
            return dt.strftime("%Y-%m-%dT%H:%M:%SZ")
        except ValueError:
            continue
    
    raise ValueError(f"Não foi possível converter data '{data_completa}' para ISO 8601")


def parsear_arquivo_texto(caminho_arquivo: Path) -> List[Dict[str, Any]]:
    """
    Parseia arquivo texto com jogos.
    
    Formato esperado:
    Rodada 1
    2024-04-13 16:00 | Flamengo x Palmeiras | Maracanã
    2024-04-13 18:30 | Corinthians x São Paulo | Neo Química Arena
    
    Args:
        caminho_arquivo: Path para o arquivo texto
        
    Returns:
        Lista de dicionários com dados dos jogos
        
    Raises:
        FileNotFoundError: Se arquivo não existir
        ValueError: Se formato do arquivo for inválido
    """
    if not caminho_arquivo.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_arquivo}")
    
    jogos = []
    rodada_atual = None
    
    try:
        with open(caminho_arquivo, 'r', encoding='utf-8') as f:
            linhas = f.readlines()
    except Exception as e:
        raise ValueError(f"Erro ao ler arquivo: {str(e)}")
    
    for i, linha in enumerate(linhas, 1):
        linha = linha.strip()
        if not linha:
            continue
        
        # Detectar rodada
        rodada_extraida = extrair_rodada(linha)
        if rodada_extraida:
            rodada_atual = rodada_extraida
            continue
        
        # Parsear linha de jogo
        # Formato: "data hora | mandante x visitante | local" ou "data hora - mandante x visitante - local"
        match = re.match(r'(.+?)\s*[\|\-]\s*(.+?)\s+x\s+(.+?)\s*[\|\-]\s*(.+)', linha, re.IGNORECASE)
        if match:
            data_hora_str = match.group(1).strip()
            mandante = match.group(2).strip()
            visitante = match.group(3).strip()
            local = match.group(4).strip()
            
            # Separar data e hora
            partes_data_hora = data_hora_str.split()
            if len(partes_data_hora) >= 2:
                data_str = partes_data_hora[0]
                hora_str = partes_data_hora[1]
            else:
                data_str = data_hora_str
                hora_str = None
            
            try:
                data_iso = converter_data_iso8601(data_str, hora_str)
            except ValueError as e:
                raise ValueError(f"Linha {i}: {str(e)}")
            
            jogo = {
                'rodada': rodada_atual,
                'mandante': mandante,
                'visitante': visitante,
                'data': data_iso,
                'local': local
            }
            
            jogos.append(jogo)
        else:
            # Tentar formato alternativo sem local
            # Formato: "data hora | mandante x visitante" ou "data hora - mandante x visitante"
            match_alt = re.match(r'(.+?)\s*[\|\-]\s*(.+?)\s+x\s+(.+)', linha, re.IGNORECASE)
            if match_alt:
                data_hora_str = match_alt.group(1).strip()
                mandante = match_alt.group(2).strip()
                visitante = match_alt.group(3).strip()
                local = "A definir"
                
                # Separar data e hora
                partes_data_hora = data_hora_str.split()
                if len(partes_data_hora) >= 2:
                    data_str = partes_data_hora[0]
                    hora_str = partes_data_hora[1]
                else:
                    data_str = data_hora_str
                    hora_str = None
                
                try:
                    data_iso = converter_data_iso8601(data_str, hora_str)
                except ValueError as e:
                    raise ValueError(f"Linha {i}: {str(e)}")
                
                jogo = {
                    'rodada': rodada_atual,
                    'mandante': mandante,
                    'visitante': visitante,
                    'data': data_iso,
                    'local': local
                }
                
                jogos.append(jogo)
            else:
                # Linha não reconhecida - pular com aviso
                print(f"Aviso: Linha {i} não reconhecida: {linha}")
    
    if not jogos:
        raise ValueError("Nenhum jogo encontrado no arquivo")
    
    return jogos


def parsear_planilha_excel(caminho_arquivo: Path) -> List[Dict[str, Any]]:
    """
    Parseia planilha Excel com jogos.
    
    Colunas esperadas: Rodada, Data, Hora, Mandante, Visitante, Local
    
    Args:
        caminho_arquivo: Path para o arquivo Excel
        
    Returns:
        Lista de dicionários com dados dos jogos
        
    Raises:
        ImportError: Se openpyxl não estiver instalado
        FileNotFoundError: Se arquivo não existir
        ValueError: Se formato da planilha for inválido
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("Biblioteca openpyxl não encontrada. Instale com: pip install openpyxl")
    
    if not caminho_arquivo.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_arquivo}")
    
    try:
        workbook = load_workbook(caminho_arquivo)
        sheet = workbook.active
    except Exception as e:
        raise ValueError(f"Erro ao abrir planilha Excel: {str(e)}")
    
    # Encontrar cabeçalhos
    headers = {}
    primeira_linha = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    
    for i, header in enumerate(primeira_linha):
        if header:
            header_lower = str(header).lower().strip()
            if 'rodada' in header_lower:
                headers['rodada'] = i
            elif 'data' in header_lower:
                headers['data'] = i
            elif 'hora' in header_lower:
                headers['hora'] = i
            elif 'mandante' in header_lower or 'casa' in header_lower:
                headers['mandante'] = i
            elif 'visitante' in header_lower or 'fora' in header_lower:
                headers['visitante'] = i
            elif 'local' in header_lower or 'estádio' in header_lower or 'estadio' in header_lower:
                headers['local'] = i
    
    # Verificar se encontrou colunas obrigatórias
    colunas_obrigatorias = ['data', 'mandante', 'visitante']
    for coluna in colunas_obrigatorias:
        if coluna not in headers:
            raise ValueError(f"Coluna obrigatória '{coluna}' não encontrada na planilha")
    
    jogos = []
    
    # Processar linhas de dados
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        if not any(row):  # Pular linhas vazias
            continue
        
        try:
            # Extrair dados da linha
            rodada = row[headers['rodada']] if 'rodada' in headers else None
            data = row[headers['data']]
            hora = row[headers['hora']] if 'hora' in headers else None
            mandante = row[headers['mandante']]
            visitante = row[headers['visitante']]
            local = row[headers['local']] if 'local' in headers else "A definir"
            
            # Validar dados obrigatórios
            if not mandante or not visitante:
                print(f"Aviso: Linha {row_num} ignorada - times não especificados")
                continue
            
            if not data:
                print(f"Aviso: Linha {row_num} ignorada - data não especificada")
                continue
            
            # Converter data
            data_str = str(data) if data else ""
            hora_str = str(hora) if hora else None
            
            # Se data é um objeto datetime do Excel
            if isinstance(data, datetime):
                data_iso = data.strftime("%Y-%m-%dT%H:%M:%SZ")
            else:
                data_iso = converter_data_iso8601(data_str, hora_str)
            
            # Converter rodada
            if rodada is not None:
                try:
                    rodada = int(rodada)
                except (ValueError, TypeError):
                    rodada = None
            
            jogo = {
                'rodada': rodada,
                'mandante': str(mandante).strip(),
                'visitante': str(visitante).strip(),
                'data': data_iso,
                'local': str(local).strip() if local else "A definir"
            }
            
            jogos.append(jogo)
            
        except Exception as e:
            raise ValueError(f"Erro na linha {row_num}: {str(e)}")
    
    if not jogos:
        raise ValueError("Nenhum jogo encontrado na planilha")
    
    return jogos


def organizar_jogos_por_rodadas(jogos: List[Dict[str, Any]]) -> Dict[int, List[Dict[str, Any]]]:
    """
    Organiza jogos por rodadas.
    
    Args:
        jogos: Lista de jogos importados
        
    Returns:
        Dicionário com jogos organizados por número da rodada
    """
    rodadas = {}
    rodada_atual = 1
    
    for jogo in jogos:
        # Se jogo tem rodada especificada, usar ela
        if jogo.get('rodada') is not None:
            num_rodada = jogo['rodada']
        else:
            # Se não tem rodada, usar rodada atual e incrementar quando necessário
            num_rodada = rodada_atual
        
        if num_rodada not in rodadas:
            rodadas[num_rodada] = []
        
        rodadas[num_rodada].append(jogo)
        
        # Atualizar rodada atual para próximo jogo sem rodada especificada
        if jogo.get('rodada') is None:
            rodada_atual = max(rodada_atual, num_rodada + 1)
    
    return rodadas


def normalizar_nomes_times(jogos: List[Dict[str, Any]], times_existentes: Optional[List[str]] = None) -> List[Dict[str, Any]]:
    """
    Normaliza nomes de times nos jogos.
    
    Args:
        jogos: Lista de jogos
        times_existentes: Lista de times já existentes na tabela (para sugestões)
        
    Returns:
        Lista de jogos com nomes normalizados
    """
    jogos_normalizados = []
    
    for jogo in jogos:
        jogo_normalizado = jogo.copy()
        
        # Normalizar mandante
        mandante_original = jogo['mandante']
        mandante_normalizado = normalizar_nome_time(mandante_original)
        
        # Tentar encontrar similar se há times existentes
        if times_existentes:
            similar = encontrar_time_similar(mandante_original, times_existentes)
            if similar:
                jogo_normalizado['mandante'] = similar
            else:
                jogo_normalizado['mandante'] = mandante_original
        else:
            jogo_normalizado['mandante'] = mandante_original
        
        # Normalizar visitante
        visitante_original = jogo['visitante']
        visitante_normalizado = normalizar_nome_time(visitante_original)
        
        # Tentar encontrar similar se há times existentes
        if times_existentes:
            similar = encontrar_time_similar(visitante_original, times_existentes)
            if similar:
                jogo_normalizado['visitante'] = similar
            else:
                jogo_normalizado['visitante'] = visitante_original
        else:
            jogo_normalizado['visitante'] = visitante_original
        
        jogos_normalizados.append(jogo_normalizado)
    
    return jogos_normalizados


def criar_estrutura_jogo(jogo: Dict[str, Any], id_jogo: str) -> Dict[str, Any]:
    """
    Cria estrutura completa de um jogo para a tabela.
    
    Args:
        jogo: Dados básicos do jogo
        id_jogo: ID único do jogo
        
    Returns:
        Dicionário com estrutura completa do jogo
    """
    return {
        "id": id_jogo,
        "mandante": jogo['mandante'],
        "visitante": jogo['visitante'],
        "data": jogo['data'],
        "local": jogo['local'],
        "gols_mandante": 0,
        "gols_visitante": 0,
        "status": "agendado",
        "obrigatorio": True
    }


def atualizar_tabela_json(caminho_tabela: Path, jogos_organizados: Dict[int, List[Dict[str, Any]]], 
                         mesclar: bool = False) -> None:
    """
    Atualiza arquivo tabela.json com os jogos importados.
    
    Args:
        caminho_tabela: Path para o arquivo tabela.json
        jogos_organizados: Jogos organizados por rodada
        mesclar: Se True, mescla com dados existentes; se False, sobrescreve
        
    Raises:
        ValueError: Se estrutura da tabela for inválida
        OSError: Se não conseguir salvar o arquivo
    """
    # Carregar tabela existente
    if caminho_tabela.exists() and mesclar:
        try:
            with open(caminho_tabela, 'r', encoding='utf-8') as f:
                tabela = json.load(f)
        except Exception as e:
            raise ValueError(f"Erro ao carregar tabela existente: {str(e)}")
    else:
        # Se não existe ou não deve mesclar, usar estrutura básica
        tabela = {
            "campeonato": "",
            "temporada": "",
            "rodada_atual": 0,
            "data_atualizacao": datetime.now().strftime("%Y-%m-%dT%H:%M:%SZ"),
            "codigo_campeonato": "",
            "rodadas": []
        }
    
    # Preparar contador de IDs
    contador_id = 1
    
    # Se mesclando, encontrar próximo ID disponível
    if mesclar and 'rodadas' in tabela:
        for rodada in tabela['rodadas']:
            if 'jogos' in rodada:
                for jogo in rodada['jogos']:
                    if 'id' in jogo and jogo['id'].startswith('jogo-'):
                        try:
                            num_id = int(jogo['id'].split('-')[1])
                            contador_id = max(contador_id, num_id + 1)
                        except (ValueError, IndexError):
                            pass
    
    # Converter jogos organizados para estrutura da tabela
    rodadas_existentes = {r['numero']: r for r in tabela.get('rodadas', [])} if mesclar else {}
    
    for num_rodada, jogos_rodada in sorted(jogos_organizados.items()):
        # Se rodada já existe e estamos mesclando, adicionar jogos à rodada existente
        if num_rodada in rodadas_existentes:
            rodada = rodadas_existentes[num_rodada]
            if 'jogos' not in rodada:
                rodada['jogos'] = []
        else:
            # Criar nova rodada
            rodada = {
                "numero": num_rodada,
                "jogos": []
            }
            rodadas_existentes[num_rodada] = rodada
        
        # Adicionar jogos à rodada
        for jogo in jogos_rodada:
            id_jogo = gerar_id_jogo(contador_id)
            contador_id += 1
            
            jogo_completo = criar_estrutura_jogo(jogo, id_jogo)
            rodada['jogos'].append(jogo_completo)
    
    # Atualizar lista de rodadas na tabela
    tabela['rodadas'] = list(rodadas_existentes.values())
    tabela['rodadas'].sort(key=lambda r: r['numero'])
    
    # Atualizar timestamp
    tabela['data_atualizacao'] = datetime.now().strftime("%Y-%m-%dT%H:%M:%SZ")
    
    # Validar estrutura antes de salvar
    valido, erros = validar_estrutura_tabela(tabela)
    if not valido:
        raise ValueError(f"Estrutura da tabela inválida após importação: {'; '.join(erros)}")
    
    # Salvar arquivo
    try:
        with open(caminho_tabela, 'w', encoding='utf-8') as f:
            json.dump(tabela, f, indent=2, ensure_ascii=False)
    except Exception as e:
        raise OSError(f"Erro ao salvar tabela: {str(e)}")


def obter_times_existentes(caminho_tabela: Path) -> List[str]:
    """
    Obtém lista de times já existentes na tabela.
    
    Args:
        caminho_tabela: Path para o arquivo tabela.json
        
    Returns:
        Lista de nomes de times únicos
    """
    if not caminho_tabela.exists():
        return []
    
    try:
        with open(caminho_tabela, 'r', encoding='utf-8') as f:
            tabela = json.load(f)
    except Exception:
        return []
    
    times = set()
    
    if 'rodadas' in tabela:
        for rodada in tabela['rodadas']:
            if 'jogos' in rodada:
                for jogo in rodada['jogos']:
                    if 'mandante' in jogo:
                        times.add(jogo['mandante'])
                    if 'visitante' in jogo:
                        times.add(jogo['visitante'])
    
    return list(times)


def confirmar_operacao(mensagem: str) -> bool:
    """
    Solicita confirmação do usuário para operações críticas.
    
    Args:
        mensagem: Mensagem a ser exibida
        
    Returns:
        True se o usuário confirmar, False caso contrário
    """
    resposta = input(f"{mensagem} (s/n): ").strip().lower()
    return resposta in ['s', 'sim', 'y', 'yes']


def validar_dados_importados(jogos: List[Dict[str, Any]]) -> Tuple[bool, List[str]]:
    """
    Valida dados importados antes de processar.
    
    Args:
        jogos: Lista de jogos importados
        
    Returns:
        Tupla (sucesso, lista_de_erros)
    """
    erros = []
    
    if not jogos:
        erros.append("Nenhum jogo foi importado")
        return False, erros
    
    for i, jogo in enumerate(jogos, 1):
        # Verificar campos obrigatórios
        campos_obrigatorios = ['mandante', 'visitante', 'data', 'local']
        for campo in campos_obrigatorios:
            if campo not in jogo or not jogo[campo]:
                erros.append(f"Jogo {i}: Campo '{campo}' ausente ou vazio")
        
        # Validar data
        if 'data' in jogo:
            valido, erro_data = validar_data(jogo['data'])
            if not valido:
                erros.append(f"Jogo {i}: {erro_data}")
        
        # Verificar se mandante e visitante são diferentes
        if 'mandante' in jogo and 'visitante' in jogo:
            if jogo['mandante'].strip().lower() == jogo['visitante'].strip().lower():
                erros.append(f"Jogo {i}: Mandante e visitante não podem ser o mesmo time")
    
    return len(erros) == 0, erros


def importar_tabela(campeonato: str, arquivo: Optional[str] = None, excel: Optional[str] = None, 
                   mesclar: bool = False, forcar: bool = False) -> bool:
    """
    Função principal para importar tabela de jogos.
    
    Args:
        campeonato: Nome do campeonato
        arquivo: Path para arquivo texto (opcional)
        excel: Path para arquivo Excel (opcional)
        mesclar: Se True, mescla com dados existentes
        forcar: Se True, não solicita confirmação
        
    Returns:
        True se importação foi bem-sucedida, False caso contrário
    """
    try:
        # Validar argumentos
        if not arquivo and not excel:
            print("Erro: Deve especificar --arquivo ou --excel")
            return False
        
        if arquivo and excel:
            print("Erro: Especifique apenas --arquivo OU --excel, não ambos")
            return False
        
        # Verificar se campeonato existe
        caminho_campeonato = CAMPEONATOS_DIR / campeonato
        if not caminho_campeonato.exists():
            print(f"Erro: Campeonato '{campeonato}' não encontrado")
            return False
        
        caminho_tabela = caminho_campeonato / "Tabela" / ARQUIVO_TABELA
        
        # Verificar se deve mesclar ou sobrescrever
        if caminho_tabela.exists() and not mesclar and not forcar:
            print(f"Aviso: Arquivo tabela.json já existe no campeonato '{campeonato}'")
            if not confirmar_operacao("Deseja sobrescrever os dados existentes?"):
                print("Operação cancelada pelo usuário")
                return False
        
        # Determinar tipo de arquivo e parsear
        if arquivo:
            caminho_arquivo = Path(arquivo)
            print(f"Importando jogos do arquivo texto: {caminho_arquivo}")
            jogos = parsear_arquivo_texto(caminho_arquivo)
        else:
            caminho_arquivo = Path(excel)
            print(f"Importando jogos da planilha Excel: {caminho_arquivo}")
            jogos = parsear_planilha_excel(caminho_arquivo)
        
        print(f"✓ {len(jogos)} jogos encontrados no arquivo")
        
        # Validar dados importados
        valido, erros = validar_dados_importados(jogos)
        if not valido:
            print("Erro: Dados importados contêm erros:")
            for erro in erros:
                print(f"  - {erro}")
            return False
        
        print("✓ Dados importados validados com sucesso")
        
        # Obter times existentes para normalização
        times_existentes = obter_times_existentes(caminho_tabela)
        if times_existentes:
            print(f"✓ {len(times_existentes)} times encontrados na tabela existente")
        
        # Normalizar nomes de times
        jogos_normalizados = normalizar_nomes_times(jogos, times_existentes)
        print("✓ Nomes de times normalizados")
        
        # Organizar por rodadas
        jogos_organizados = organizar_jogos_por_rodadas(jogos_normalizados)
        total_rodadas = len(jogos_organizados)
        print(f"✓ Jogos organizados em {total_rodadas} rodadas")
        
        # Mostrar resumo
        print("\n📊 Resumo da importação:")
        for num_rodada in sorted(jogos_organizados.keys()):
            qtd_jogos = len(jogos_organizados[num_rodada])
            print(f"  Rodada {num_rodada}: {qtd_jogos} jogos")
        
        # Confirmar antes de salvar (se não forçado)
        if not forcar:
            acao = "mesclar com" if mesclar else "sobrescrever"
            if not confirmar_operacao(f"Confirma importação? Isso irá {acao} a tabela existente"):
                print("Operação cancelada pelo usuário")
                return False
        
        # Atualizar tabela.json
        atualizar_tabela_json(caminho_tabela, jogos_organizados, mesclar)
        print("✓ Arquivo tabela.json atualizado com sucesso")
        
        print(f"\n🎉 Importação concluída com sucesso!")
        print(f"📁 Arquivo: {caminho_tabela}")
        print(f"📊 Total: {len(jogos)} jogos em {total_rodadas} rodadas")
        
        # Sugerir próximos passos
        print("\n📋 Próximos passos sugeridos:")
        print(f"1. Verifique o arquivo tabela.json gerado")
        print(f"2. Execute: python importar_palpites.py --campeonato '{campeonato}' --arquivo palpites.txt")
        
        return True
        
    except FileNotFoundError as e:
        print(f"Erro: Arquivo não encontrado - {str(e)}")
        return False
    except ValueError as e:
        print(f"Erro de validação: {str(e)}")
        return False
    except OSError as e:
        print(f"Erro de sistema: {str(e)}")
        return False
    except Exception as e:
        print(f"Erro inesperado: {str(e)}")
        return False


def main():
    """Função principal do script."""
    parser = argparse.ArgumentParser(
        description="Importa tabela de jogos de arquivo externo para tabela.json",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos de uso:
  python importar_tabela.py --campeonato "Brasileirao-2025" --arquivo "jogos.txt"
  python importar_tabela.py --campeonato "Copa-Brasil" --excel "tabela.xlsx"
  python importar_tabela.py --campeonato "Paulistao" --arquivo "jogos.txt" --mesclar
  python importar_tabela.py --campeonato "Libertadores" --excel "jogos.xlsx" --forcar

Formato do arquivo texto:
  Rodada 1
  2024-04-13 16:00 | Flamengo x Palmeiras | Maracanã
  2024-04-13 18:30 | Corinthians x São Paulo | Neo Química Arena
  
  Rodada 2
  2024-04-20 16:00 | Palmeiras x Corinthians | Allianz Parque

Formato da planilha Excel:
  Colunas: Rodada, Data, Hora, Mandante, Visitante, Local
  (Colunas Rodada, Hora e Local são opcionais)
        """
    )
    
    parser.add_argument(
        '--campeonato',
        required=True,
        help='Nome do campeonato (deve existir)'
    )
    
    parser.add_argument(
        '--arquivo',
        help='Arquivo texto com jogos'
    )
    
    parser.add_argument(
        '--excel',
        help='Planilha Excel com jogos'
    )
    
    parser.add_argument(
        '--mesclar',
        action='store_true',
        help='Mescla com dados existentes ao invés de sobrescrever'
    )
    
    parser.add_argument(
        '--forcar',
        action='store_true',
        help='Força importação sem confirmação'
    )
    
    args = parser.parse_args()
    
    # Validar argumentos
    if not args.campeonato.strip():
        print("Erro: Nome do campeonato não pode estar vazio")
        sys.exit(1)
    
    # Executar importação
    sucesso = importar_tabela(
        campeonato=args.campeonato.strip(),
        arquivo=args.arquivo,
        excel=args.excel,
        mesclar=args.mesclar,
        forcar=args.forcar
    )
    
    if sucesso:
        print("\n✅ Script executado com sucesso!")
        sys.exit(0)
    else:
        print("\n❌ Falha na execução do script")
        sys.exit(1)


if __name__ == "__main__":
    main()